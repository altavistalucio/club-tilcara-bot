# ============================================================
#  SIMULADOR DE CHATBOT — Club Tilcara
#  Proceso: Solicitud de Licencia de Jugador
#  TPI — Organización Empresarial — UTN TUPaD 2026
#  Integrantes: Lucio y Lucía
# ============================================================
#
#  Cómo usar:
#    1. Asegurate de tener el archivo
#       ClubTilcara_Planilla_Licencias.xlsx en la misma carpeta.
#    2. Ejecutar con: python simulador_bot.py
#
# ============================================================

import openpyxl
from datetime import date

ARCHIVO_EXCEL = "ClubTilcara_Planilla_Licencias.xlsx"

# ── SEPARADOR VISUAL ─────────────────────────────────────
def linea():
    print("-" * 50)

# ── MENSAJES DEL BOT ─────────────────────────────────────
def bot(mensaje):
    print(f"\n  🤖 BOT: {mensaje}")

def usuario_input(prompt):
    return input(f"\n  👤 VOS: {prompt} > ").strip()

# ── CARGAR PLANILLA ──────────────────────────────────────
def cargar_planilla():
    """Carga la hoja Jugadores desde el Excel y devuelve
    un diccionario con ID como clave."""
    try:
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        hoja = wb["Jugadores"]
    except FileNotFoundError:
        print(f"\nERROR: No se encontró el archivo '{ARCHIVO_EXCEL}'.")
        print("Asegurate de tenerlo en la misma carpeta que este script.")
        exit()

    jugadores = {}
    for fila in hoja.iter_rows(min_row=3, values_only=True):
        # Fila 1 y 2 son encabezados; datos desde fila 3
        id_jug = fila[0]
        if id_jug is None or not str(id_jug).isdigit():
            continue
        jugadores[int(id_jug)] = {
            "id":          int(fila[0]),
            "nombre":      fila[1],
            "dni":         fila[2],
            "posicion":    fila[3],
            "categoria":   fila[4],
            "disponibles": int(fila[5]) if fila[5] is not None else 0,
            "usados":      int(fila[6]) if fila[6] is not None else 0,
            "estado":      fila[8] if fila[8] else "Activo",
        }
    return jugadores, wb

# ── REGISTRAR LICENCIA EN EXCEL ──────────────────────────
def registrar_licencia(wb, jugador, dias_solicitados, motivo):
    """Actualiza días en hoja Jugadores y agrega registro
    en hoja Historial Licencias."""

    # -- Actualizar hoja Jugadores --
    hoja_jug = wb["Jugadores"]
    for fila in hoja_jug.iter_rows(min_row=3):
        if fila[0].value == jugador["id"]:
            fila[5].value = jugador["disponibles"] - dias_solicitados
            fila[6].value = jugador["usados"] + dias_solicitados
            fila[8].value = "En licencia"
            break

    # -- Agregar en hoja Historial Licencias --
    hoja_hist = wb["Historial Licencias"]
    ultimo_id = 1
    for fila in hoja_hist.iter_rows(min_row=3, values_only=True):
        if fila[0] is not None and str(fila[0]).isdigit():
            ultimo_id = int(fila[0]) + 1

    nueva_fila = [
        ultimo_id,
        jugador["id"],
        jugador["nombre"],
        str(date.today()),
        dias_solicitados,
        motivo,
        "Aprobado",
        str(date.today()),
        "Registrado por el bot",
    ]
    hoja_hist.append(nueva_fila)
    wb.save(ARCHIVO_EXCEL)

# ════════════════════════════════════════════════════════
#  MÁQUINA DE ESTADOS
# ════════════════════════════════════════════════════════

def estado_inicio():
    """ESTADO: INICIO — saludo y solicitud de ID."""
    linea()
    bot("¡Bienvenido al sistema de licencias del Club Tilcara!")
    bot("Podés escribir 'salir' en cualquier momento para cancelar.")
    bot("¿Cuál es tu número de ID de jugador?")

def estado_validar_jugador(jugadores):
    """ESTADO: VALIDAR_JUGADOR — verifica si el ID existe."""
    while True:
        entrada = usuario_input("Ingresá tu ID")

        if entrada.lower() == "salir":
            return None, "salir"

        # Camino infeliz: no es número
        if not entrada.isdigit():
            bot("⚠️  Por favor ingresá solo números para el ID. Intentá de nuevo.")
            continue

        id_jug = int(entrada)

        # Camino infeliz: ID inexistente
        if id_jug not in jugadores:
            bot(f"⚠️  No encontré ningún jugador con el ID {id_jug}.")
            bot("Verificá el número e intentá de nuevo.")
            continue

        jugador = jugadores[id_jug]

        # Camino infeliz: suspendido
        if jugador["estado"] == "Suspendido":
            bot(f"⚠️  Hola, {jugador['nombre']}.")
            bot("Tu cuenta está suspendida. Contactá al club para más información.")
            return None, "suspendido"

        bot(f"✅ Jugador encontrado: {jugador['nombre']} — {jugador['categoria']}")
        bot(f"   Días disponibles: {jugador['disponibles']} | Días usados: {jugador['usados']}")
        return jugador, "ok"

def estado_consultar_dias(jugador):
    """ESTADO: CONSULTAR_DIAS — verifica si tiene días."""
    if jugador["disponibles"] <= 0:
        bot(f"❌ Lo sentimos, {jugador['nombre']}.")
        bot("No tenés días de licencia disponibles.")
        bot("Contactá al Director Técnico si creés que es un error.")
        return False
    return True

def estado_pedir_motivo():
    """ESTADO: PEDIR_MOTIVO — solicita el motivo."""
    bot("¿Cuál es el motivo de tu solicitud de licencia?")
    while True:
        motivo = usuario_input("Motivo")

        if motivo.lower() == "salir":
            return None

        # Camino infeliz: motivo vacío o muy corto
        if len(motivo) < 3:
            bot("⚠️  Por favor describí brevemente el motivo (mínimo 3 caracteres).")
            continue

        return motivo

def estado_pedir_dias(jugador):
    """ESTADO: PEDIR_CANTIDAD_DIAS — solicita cuántos días."""
    bot(f"¿Cuántos días necesitás? (Tenés {jugador['disponibles']} disponibles)")
    while True:
        entrada = usuario_input("Cantidad de días")

        if entrada.lower() == "salir":
            return None

        # Camino infeliz: no es número
        if not entrada.isdigit():
            bot("⚠️  Por favor ingresá la cantidad como número (ej: 3).")
            continue

        dias = int(entrada)

        # Camino infeliz: cero o negativo
        if dias <= 0:
            bot("⚠️  La cantidad de días debe ser al menos 1.")
            continue

        # Camino infeliz: supera los disponibles
        if dias > jugador["disponibles"]:
            bot(f"⚠️  Solo tenés {jugador['disponibles']} días disponibles.")
            bot(f"   Podés solicitar hasta ese máximo.")
            continue

        return dias

def estado_esperar_dt(jugador, dias, motivo):
    """ESTADO: ESPERAR_DT — simula la decisión del DT."""
    linea()
    bot("📨 Solicitud enviada al Director Técnico.")
    bot(f"   Jugador:  {jugador['nombre']}")
    bot(f"   Días:     {dias}")
    bot(f"   Motivo:   {motivo}")
    linea()

    # Simulación: el DT decide (en un sistema real vendría de otra interfaz)
    bot("(Simulación) El Director Técnico debe ingresar su decisión:")
    while True:
        decision = usuario_input("DT — Escribí 'aprobar' o 'rechazar'").lower()

        if decision == "aprobar":
            return True
        elif decision == "rechazar":
            return False
        else:
            bot("⚠️  Opción no válida. Escribí 'aprobar' o 'rechazar'.")

def estado_aprobar(wb, jugador, dias, motivo):
    """ESTADO: APROBAR — registra y notifica."""
    registrar_licencia(wb, jugador, dias, motivo)
    linea()
    bot(f"✅ ¡Licencia APROBADA para {jugador['nombre']}!")
    bot(f"   Días otorgados: {dias}")
    bot(f"   Días restantes: {jugador['disponibles'] - dias}")
    bot("   La planilla fue actualizada correctamente.")

def estado_rechazar_dt(jugador):
    """ESTADO: RECHAZAR_DT — notifica rechazo del DT."""
    linea()
    bot(f"❌ Lo sentimos, {jugador['nombre']}.")
    bot("El Director Técnico rechazó tu solicitud.")
    bot("Podés comunicarte con él para más detalles.")

# ════════════════════════════════════════════════════════
#  FLUJO PRINCIPAL
# ════════════════════════════════════════════════════════

def ejecutar_bot():
    jugadores, wb = cargar_planilla()

    while True:
        # INICIO
        estado_inicio()

        # VALIDAR JUGADOR
        jugador, resultado = estado_validar_jugador(jugadores)
        if resultado in ("salir", "suspendido"):
            if resultado == "salir":
                bot("Proceso cancelado. ¡Hasta luego!")
            linea()
            break

        # CONSULTAR DÍAS
        tiene_dias = estado_consultar_dias(jugador)
        if not tiene_dias:
            linea()
            reintentar = usuario_input("¿Querés iniciar una nueva consulta? (si/no)")
            if reintentar.lower() == "si":
                continue
            else:
                bot("¡Hasta luego!")
                linea()
                break

        # PEDIR MOTIVO
        motivo = estado_pedir_motivo()
        if motivo is None:
            bot("Proceso cancelado. ¡Hasta luego!")
            linea()
            break

        # PEDIR CANTIDAD DE DÍAS
        dias = estado_pedir_dias(jugador)
        if dias is None:
            bot("Proceso cancelado. ¡Hasta luego!")
            linea()
            break

        # ESPERAR DT
        aprobado = estado_esperar_dt(jugador, dias, motivo)

        # RESULTADO
        if aprobado:
            estado_aprobar(wb, jugador, dias, motivo)
        else:
            estado_rechazar_dt(jugador)

        linea()
        reintentar = usuario_input("¿Querés realizar otra consulta? (si/no)")
        if reintentar.lower() != "si":
            bot("¡Hasta luego!")
            linea()
            break

# ── PUNTO DE ENTRADA ─────────────────────────────────────
if __name__ == "__main__":
    ejecutar_bot()
